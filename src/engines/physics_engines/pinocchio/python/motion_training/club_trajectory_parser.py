"""Parser for club trajectory data from motion capture Excel files.

This module parses golf club motion capture data typically stored in Excel format
with position and orientation data for grip (mid-hands) and club face markers.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np

if TYPE_CHECKING:
    from numpy.typing import NDArray

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class SwingEventMarkers:
    """Markers for key swing events (frame indices)."""

    address: int = 0  # A - Address/setup position
    top: int = 0  # T - Top of backswing
    impact: int = 0  # I - Ball impact
    finish: int = 0  # F - End of follow-through
    club_head_speed: float = 0.0  # CHS in mph


@dataclass
class ClubFrame:
    """Single frame of club motion data."""

    time: float  # Time in seconds (relative to impact)
    sample_index: int  # Frame index

    # Grip (mid-hands) position and orientation
    grip_position: NDArray[np.float64]  # [x, y, z] in meters
    grip_rotation: NDArray[np.float64]  # 3x3 rotation matrix

    # Club face position and orientation
    club_face_position: NDArray[np.float64]  # [x, y, z] in meters
    club_face_rotation: NDArray[np.float64]  # 3x3 rotation matrix


@dataclass
class ClubTrajectory:
    """Complete club trajectory with all frames and metadata."""

    frames: list[ClubFrame] = field(default_factory=list)
    events: SwingEventMarkers = field(default_factory=SwingEventMarkers)
    sample_rate_hz: float = 240.0  # Default to 240 Hz motion capture

    @property
    def num_frames(self) -> int:
        """Return number of frames in trajectory."""
        return len(self.frames)

    @property
    def duration(self) -> float:
        """Return total duration in seconds."""
        if not self.frames:
            return 0.0
        return self.frames[-1].time - self.frames[0].time

    @property
    def times(self) -> NDArray[np.float64]:
        """Return array of all timestamps."""
        return np.array([f.time for f in self.frames])

    @property
    def grip_positions(self) -> NDArray[np.float64]:
        """Return Nx3 array of grip positions."""
        return np.array([f.grip_position for f in self.frames])

    @property
    def club_face_positions(self) -> NDArray[np.float64]:
        """Return Nx3 array of club face positions."""
        return np.array([f.club_face_position for f in self.frames])

    def get_frame_at_time(self, t: float) -> ClubFrame:
        """Interpolate to get frame at specific time."""
        times = self.times
        if t <= times[0]:
            return self.frames[0]
        if t >= times[-1]:
            return self.frames[-1]

        # Find bracketing frames
        idx = np.searchsorted(times, t)
        t0, t1 = times[idx - 1], times[idx]
        alpha = (t - t0) / (t1 - t0)

        f0, f1 = self.frames[idx - 1], self.frames[idx]

        # Linear interpolation for positions
        grip_pos = (1 - alpha) * f0.grip_position + alpha * f1.grip_position
        face_pos = (1 - alpha) * f0.club_face_position + alpha * f1.club_face_position

        # SLERP-like interpolation for rotations (simplified)
        # Note: Using nearest-neighbor for rotation; SLERP would improve this
        grip_rot = f0.grip_rotation
        face_rot = f0.club_face_rotation

        return ClubFrame(
            time=t,
            sample_index=int((idx - 1 + alpha) * 100),
            grip_position=grip_pos,
            grip_rotation=grip_rot,
            club_face_position=face_pos,
            club_face_rotation=face_rot,
        )

    def get_event_frame(self, event: str) -> ClubFrame | None:
        """Get frame at a specific swing event (address, top, impact, finish)."""
        event_map = {
            "address": self.events.address,
            "top": self.events.top,
            "impact": self.events.impact,
            "finish": self.events.finish,
        }
        if event not in event_map:
            return None

        frame_idx = event_map[event]
        # Find the frame with matching sample index
        for frame in self.frames:
            if frame.sample_index == frame_idx:
                return frame
        return None


class ClubTrajectoryParser:
    """Parser for club trajectory data from Excel files.

    Expected Excel format:
    - Row 0: Metadata (ball type, event markers A=, T=, I=, F=, CHS=)
    - Row 1: Marker names (Mid-hands, Center of club face)
    - Row 2: Column headers (Sample #, Time, X, Y, Z, Xx, Xy, Xz, Yx, Yy, Yz, Zx,
      Zy, Zz)
    - Row 3+: Data

    Positions are in centimeters and converted to meters.
    """

    # Column indices for Mid-hands data
    SAMPLE_COL = 0
    TIME_COL = 1
    GRIP_X_COL = 2
    GRIP_Y_COL = 3
    GRIP_Z_COL = 4
    GRIP_XX_COL = 5
    GRIP_XY_COL = 6
    GRIP_XZ_COL = 7
    GRIP_YX_COL = 8
    GRIP_YY_COL = 9
    GRIP_YZ_COL = 10

    # Column indices for Club face data
    FACE_X_COL = 14
    FACE_Y_COL = 15
    FACE_Z_COL = 16
    FACE_XX_COL = 17
    FACE_XY_COL = 18
    FACE_XZ_COL = 19
    FACE_YX_COL = 20
    FACE_YY_COL = 21
    FACE_YZ_COL = 22

    # Conversion factor from centimeters to meters
    CM_TO_M = 0.01

    def __init__(self, file_path: str | Path) -> None:
        """Initialize parser with file path.

        Args:
            file_path: Path to the Excel file containing club trajectory data
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")

    def parse(self, sheet_name: str = "TW_wiffle") -> ClubTrajectory:
        """Parse the Excel file and return ClubTrajectory.

        Args:
            sheet_name: Name of the sheet to parse

        Returns:
            ClubTrajectory object with all frames and metadata
        """
        if PANDAS_AVAILABLE:
            return self._parse_with_pandas(sheet_name)
        elif OPENPYXL_AVAILABLE:
            return self._parse_with_openpyxl(sheet_name)
        else:
            raise ImportError(
                "Either pandas or openpyxl is required to parse Excel files. "
                "Install with: pip install pandas openpyxl"
            )

    def _parse_with_pandas(self, sheet_name: str) -> ClubTrajectory:
        """Parse using pandas."""
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)

        # Parse event markers from row 0
        events = self._parse_events_pandas(df.iloc[0])

        # Parse data starting from row 3 (skip metadata, marker names, headers)
        trajectory = ClubTrajectory(events=events)

        for idx in range(3, len(df)):
            row = df.iloc[idx]
            frame = self._parse_row(row)
            if frame is not None:
                trajectory.frames.append(frame)

        return trajectory

    def _parse_with_openpyxl(self, sheet_name: str) -> ClubTrajectory:
        """Parse using openpyxl."""
        wb = load_workbook(self.file_path, data_only=True)
        sheet = wb[sheet_name]

        # Parse event markers from row 1
        row1 = [cell.value for cell in sheet[1]]
        events = self._parse_events_list(row1)

        # Parse data starting from row 4 (1-indexed, skip metadata rows)
        trajectory = ClubTrajectory(events=events)

        for row_idx in range(4, sheet.max_row + 1):
            row = [cell.value for cell in sheet[row_idx]]
            frame = self._parse_row(row)
            if frame is not None:
                trajectory.frames.append(frame)

        wb.close()
        return trajectory

    def _parse_events_pandas(self, row: pd.Series) -> SwingEventMarkers:
        """Parse event markers from pandas row."""
        return self._parse_events_list(row.tolist())

    def _parse_events_list(self, row: list) -> SwingEventMarkers:
        """Parse event markers from list."""
        events = SwingEventMarkers()

        for i, val in enumerate(row):
            if val == "A=" and i + 1 < len(row):
                events.address = int(row[i + 1]) if row[i + 1] else 0
            elif val == "T=" and i + 1 < len(row):
                events.top = int(row[i + 1]) if row[i + 1] else 0
            elif val == "I=" and i + 1 < len(row):
                events.impact = int(row[i + 1]) if row[i + 1] else 0
            elif val == "F=" and i + 1 < len(row):
                events.finish = int(row[i + 1]) if row[i + 1] else 0
            elif val == "CHS" and i + 1 < len(row):
                events.club_head_speed = float(row[i + 1]) if row[i + 1] else 0.0

        return events

    @staticmethod
    def _make_row_accessor(row):
        if hasattr(row, "iloc"):

            def get(i) -> object | None:
                """Return the value at index i from a pandas row."""
                return row.iloc[i] if i < len(row) else None

        else:

            def get(i) -> object | None:
                """Return the value at index i from a list row."""
                return row[i] if i < len(row) else None

        return get

    @staticmethod
    def _orthogonalize_axes(x_axis, y_axis):
        x_axis = x_axis / (np.linalg.norm(x_axis) + 1e-8)
        y_axis = y_axis - np.dot(y_axis, x_axis) * x_axis
        y_axis = y_axis / (np.linalg.norm(y_axis) + 1e-8)
        z_axis = np.cross(x_axis, y_axis)
        return np.column_stack([x_axis, y_axis, z_axis])

    def _parse_grip_data(self, get):
        grip_pos = np.array(
            [
                float(get(self.GRIP_X_COL)) * self.CM_TO_M,
                float(get(self.GRIP_Y_COL)) * self.CM_TO_M,
                float(get(self.GRIP_Z_COL)) * self.CM_TO_M,
            ]
        )

        grip_x = np.array(
            [
                float(get(self.GRIP_XX_COL) or 1.0),
                float(get(self.GRIP_XY_COL) or 0.0),
                float(get(self.GRIP_XZ_COL) or 0.0),
            ]
        )
        grip_y = np.array(
            [
                float(get(self.GRIP_YX_COL) or 0.0),
                float(get(self.GRIP_YY_COL) or 1.0),
                float(get(self.GRIP_YZ_COL) or 0.0),
            ]
        )
        grip_rot = self._orthogonalize_axes(grip_x, grip_y)

        return grip_pos, grip_rot

    def _parse_club_face_data(self, get, grip_pos, grip_rot):
        face_x = get(self.FACE_X_COL)
        face_y = get(self.FACE_Y_COL)
        face_z = get(self.FACE_Z_COL)

        if face_x is None or face_y is None or face_z is None:
            face_pos = grip_pos + np.array([0.0, 0.0, -1.0])
            return face_pos, grip_rot.copy()

        face_pos = np.array(
            [
                float(face_x) * self.CM_TO_M,
                float(face_y) * self.CM_TO_M,
                float(face_z) * self.CM_TO_M,
            ]
        )

        face_x_axis = np.array(
            [
                float(get(self.FACE_XX_COL) or 1.0),
                float(get(self.FACE_XY_COL) or 0.0),
                float(get(self.FACE_XZ_COL) or 0.0),
            ]
        )
        face_y_axis = np.array(
            [
                float(get(self.FACE_YX_COL) or 0.0),
                float(get(self.FACE_YY_COL) or 1.0),
                float(get(self.FACE_YZ_COL) or 0.0),
            ]
        )
        face_rot = self._orthogonalize_axes(face_x_axis, face_y_axis)

        return face_pos, face_rot

    def _parse_row(self, row) -> ClubFrame | None:
        """Parse a single data row into ClubFrame."""
        get = self._make_row_accessor(row)

        sample = get(self.SAMPLE_COL)
        if sample is None or not isinstance(sample, (int, float)):
            return None

        try:
            sample_idx = int(sample)
            time = float(get(self.TIME_COL))

            grip_pos, grip_rot = self._parse_grip_data(get)
            face_pos, face_rot = self._parse_club_face_data(get, grip_pos, grip_rot)

            return ClubFrame(
                time=time,
                sample_index=sample_idx,
                grip_position=grip_pos,
                grip_rotation=grip_rot,
                club_face_position=face_pos,
                club_face_rotation=face_rot,
            )

        except (ValueError, TypeError):
            return None

    def get_available_sheets(self) -> list[str]:
        """Return list of available sheet names in the Excel file."""
        if PANDAS_AVAILABLE:
            xls = pd.ExcelFile(self.file_path)
            return xls.sheet_names
        elif OPENPYXL_AVAILABLE:
            wb = load_workbook(self.file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        else:
            return []


def compute_hand_positions(
    frame: ClubFrame,
    left_offset: float = 0.04,
    right_offset: float = -0.04,
) -> tuple[NDArray[np.float64], NDArray[np.float64]]:
    """Compute left and right hand positions from grip frame.

    Args:
        frame: Club frame with grip position and rotation
        left_offset: Offset along grip Z-axis for left hand (meters)
        right_offset: Offset along grip Z-axis for right hand (meters)

    Returns:
        Tuple of (left_hand_position, right_hand_position)
    """
    # Get the grip Z-axis (along the shaft)
    grip_z = frame.grip_rotation[:, 2]

    # Compute hand positions
    left_pos = frame.grip_position + left_offset * grip_z
    right_pos = frame.grip_position + right_offset * grip_z

    return left_pos, right_pos
